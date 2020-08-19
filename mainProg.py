##########################################################
#####      Imports                       #################
##########################################################
import tkinter as tk
from tkinter import ttk
from yahoo_fin import stock_info as si
import yfinance as yf
from tkinter import filedialog
from xlsxwriter import Workbook
from tkinter import messagebox
from selenium import webdriver as wb
from selenium.common import exceptions
import xlrd
import openpyxl
from openpyxl.styles import PatternFill
import csv
import time
import os
import shutil
import datetime
import getpass

##########################################################
#####      Constants                     #################
##########################################################

REVENUE = 3
EPS = 8
BOOKVALUE = 12
FREECASHFLOW = 15
OPERATINGCASH = 13
ROIC = 38
numbers = {}
isSaveNeeded=True
stockListFile = None
DataFileName = ".data"
numberOfSettingBrowse = 2
settingBrowseNames = ["filesLocation","favStocksLocation"]
CurrentAAABondYield = 2.41
globLabel = []
firstRow = 16
maxFiles = 250
minFiles = 200



###########################################################
###############        Page Class         #################
###########################################################
class Page(tk.Frame):
    def __init__(self,*args,**kwargs):
        tk.Frame.__init__(self,*args,**kwargs)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=18)
        self.rowconfigure(2, weight=1)
        self.grid_columnconfigure(0,weight=1)
        self.header = tk.Frame(self)
        self.content = tk.Frame(self)
        self.footer = tk.Frame(self)
        head = tk.Label(self.header, text="StockAnalysis")
        foot = tk.Label(self.footer, text="Copyright BS", anchor='s')
        head.pack()
        foot.pack()
        self.header.grid(row=0)
        self.content.grid(row=1)
        self.footer.grid(row=2)
    def show(self):
        self.pack(fill=tk.BOTH, expand=1)
    def hide(self):
        self.pack_forget()

##########################################################
#####      Frames and Widgets:           #################
##########################################################

if(os.path.isfile(DataFileName)==False):
    f = open(DataFileName,'w+')
    f.close()
root = tk.Tk()
#root.title('Title')
#root.iconbitmap('pathToImage')
root.geometry('1200x600')
root.grid_columnconfigure(0,weight=1)

##########################################################
#####      PagesDeclarations             #################
##########################################################

mainPage = Page(root)
stockDataPage = Page(root)
calValuePage = Page(root)
settingsPage = Page(root)
aboutPage = Page(root)
autoPage = Page(root)

######################################################################################
###############                 functions:                           #################
######################################################################################

def createDataFile():
    f = open(DataFileName, 'w+')
    f.close()

def rule1Calculator(epsGrowth,PE,currentEPS):
    if(epsGrowth<0 or PE<0 or currentEPS<0):
        return 0
    return PE*currentEPS*pow(epsGrowth,10)/4

def benGrahamOriginal(epsGrowth,currentEPS):
    if (epsGrowth < 0 or currentEPS < 0):
        return 0
    return (currentEPS * (8.5 + 2 * epsGrowth) * 4.4) / CurrentAAABondYield

def benGrahamUpdate(epsGrowth,currentEPS):
    if(epsGrowth<0 or currentEPS < 0):
        return 0
    return (currentEPS*(7+epsGrowth)*4.4)/CurrentAAABondYield

def getDataLocation():
    if(os.path.isfile(DataFileName)==False):
        createDataFile()
        return None
    ans = {settingBrowseNames[0]:'',settingBrowseNames[1]:''}
    i = 0
    dataFile = open(DataFileName,'r')
    for line in dataFile:
        line = line.replace("/","\\")
        ans[settingBrowseNames[i]] = line.rstrip()
        i += 1
        if i>= numberOfSettingBrowse:
            break
    dataFile.close()
    return ans

def destroyTempLabel():
    for lab in globLabel:
        lab.destroy()

def secToMonths(sec):
    return sec/60/60/24/30

def delMechanism():
    loc = getDataLocation()
    copiesLoc = loc[settingBrowseNames[1]] + "\\excel_copies"
    files = ([name for name in os.listdir(copiesLoc) if os.path.isfile(os.path.join(copiesLoc, name))])
    files.sort()
    if len(files) >= maxFiles:
        while len(files) > minFiles: # maxFiles should be around 250. minFile should be around 200
            f = files[0]
            fToDel = os.path.join(copiesLoc,f)
            files.remove(files[0])
            try:
                os.remove(fToDel)
            except:
                messagebox.showerror("ERROR!","File doesn't exist!!")

def isDataLocValid():
    dataFile = open(DataFileName,'r')
    numOfLines = 0
    for line in dataFile:
        line = line.replace("/","\\")
        numOfLines += 1
        if(os.path.isdir(line.rstrip())):
            continue
        return False
    if numOfLines < 2:
        return False
    return True

def writeFromData():
    """ a function that print to the entry the locations that are in the .data file"""
    loc = getDataLocation()
    if(loc==None):
        return
    if(len(loc) == numberOfSettingBrowse):
        stockFilesEntry.delete(0,'end')
        stockFilesEntry.insert(0, loc[settingBrowseNames[0]])
        favStocksEntry.delete(0,'end')
        favStocksEntry.insert(0, loc[settingBrowseNames[1]])
    elif(len(loc)==1):
        stockFilesEntry.delete(0, 'end')
        stockFilesEntry.insert(0, loc[settingBrowseNames[0]])


def moveLocation(oldLoc,newLoc):
    src = oldLoc[1]+"inv.xlsx"
    dest = newLoc[1]+"inv.xlsx"
    if(os.path.isfile(src)==True and os.path.isfile(dest)==False):
        shutil.move(src,dest)



def saveLocToDataFile(locList):
    for i,loc in enumerate(locList,1):
        if((not os.path.isdir(loc.rstrip())) and not loc.rstrip()==""):
            raise OSError("directory of the entry number " + str(i) + " does not exist")
    fileToWrite = open(DataFileName,'w')
    for loc in locList:
        loc = loc.rstrip() + '\n'
        fileToWrite.write(loc)
    fileToWrite.close()

def saveLocToDataFileWrapper(locList):
    try:
        saveLocToDataFile(locList)
    except OSError as err:
        messagebox.showerror("ERROR",err)
        return
    saveButton["state"]=tk.DISABLED

def browseLocation(typeOfData):
    """ a function that write to the .data file the locations of the files"""
    dir = filedialog.askdirectory()
    if (not dir == ''):
        dir += '\n'
    else:
        return
    if(os.path.isfile(DataFileName) == False):
        f = open(DataFileName, 'w+')
        f.write(dir)
    else:
        fileToRead = open(DataFileName,'r')
        numOfLines = 0
        for l in fileToRead:
            numOfLines = numOfLines+1
        lines = []
        fileToRead.close()
        fileToRead = open(DataFileName,'r')
        for line in fileToRead:
            lines.append(line)
        fileToRead.close()
        newLines = []
        for count in range(numOfLines):
            newLines.append(lines[count])
        for count in range(numOfLines,numberOfSettingBrowse):
            newLines.append("\n")
        if(typeOfData==settingBrowseNames[0]):
            newLines[0] = dir
        elif(typeOfData==settingBrowseNames[1]):
            newLines[1] = dir
        try:
            saveLocToDataFile(newLines)
        except OSError as err:
            messagebox.showerror("ERROR", err)
            return
        if(numOfLines>=2):
            moveLocation(lines,newLines)
    writeFromData()



def switchFrames(src,dest):
    src.hide()
    dest.show()

def checkCalculatorInput(inputs):
    for inp in inputs:
        try:
            float(inp)
        except:
            return False
    return True

def onFocusEntry(event):
    """a function that gets called whenever entry is clicked"""
    if event.widget.cget('fg')=='grey':
        event.widget.delete(0,"end")
        event.widget.insert(0,'')
        event.widget.config(fg='black')

def onFocusOut(event,msg):
    if(event.widget.get()==''):
        event.widget.insert(0,msg)
        event.widget.config(fg='grey')

def onChangeLoc(event):
    ans = getDataLocation()
    isChanged = False
    if not ans[settingBrowseNames[0]]==stockFilesEntry.get():
        isChanged=True
    if not ans[settingBrowseNames[1]]==favStocksEntry.get():
        isChanged=True
    if isChanged:
        saveButton["state"]=tk.NORMAL

def isCellEmpty(sheet,row,col):
    return sheet.cell_value(row,col) == ""

def isCellNegative(sheet,row,col):
    val = sheet.cell_value(row,col)
    if val=="":
        return False
    return float(sheet.cell_value(row,col)) <= 0

def checkWhichEmpty(sheet,row):
    empty = []
    for i in range(1,11):
        if isCellEmpty(sheet,row,i):
            empty.append(i)
    return empty

def checkWhichNegative(sheet,row):
    negative = []
    for i in range(1,11):
        if isCellNegative(sheet,row,i):
            negative.append(i)
    return negative

def createList(sheet,row):
    numList = []
    empties = checkWhichEmpty(sheet,row)
    #negatives = checkWhichNegative(sheet,row)
    if(len(empties)>3):
        return "Not enough Data!!"
    for cell in range(11):
        numList.append(sheet.cell_value(row, cell+1))
    return numList

def calGrowth(start,end,years,rnd):
    if(type(start)==str or type(end)==str):
        return "missing Data"
    if(start<=0 and end>0):
        return 9.99
    try:
        ans = round((pow(end/start,1/years)-1),rnd)
    except TypeError:
        return -9.99
    except ZeroDivisionError:
        return 9.99
    return ans

def findrow(filePath,stock):
    wb = xlrd.open_workbook(filePath)
    sheet = wb.sheet_by_index(0)
    i = 0
    try:
        for i in range(300):
            if not sheet.cell(i+firstRow,1):
                return i+firstRow,False
            elif sheet.cell_value(i+firstRow,1) == stock:
                return i+firstRow,True
    except:
        return i+firstRow,False

def calAverage(numbers,rnd):
    sum = 0
    numOfEmpty = 0
    for num in numbers:
        if(type(num)==str):
            numOfEmpty += 1
            continue
        sum += num
    if(numOfEmpty/len(numbers)>0.2):
        return "missing Data"
    return round((sum/(len(numbers)-numOfEmpty))/100,rnd)

def saveStock(filePath, symbol, num, color):
    copyInvFile()
    mainWB = openpyxl.load_workbook(filePath)
    ws = mainWB.active
    saveRow,isIntValueExist = findrow(filePath,symbol.upper())
    saveRow += 1
    isCash = 1
    if('FreeCashFlow' not in num.keys()):
        isCash = 0
    redFill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    greenFill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    if color == "R":
        if not isIntValueExist:
            try:
                ws.cell(column=1, row=saveRow, value=yf.Ticker(symbol).info['longName']).fill = redFill
            except:
                pass
            ws.cell(column=18, row=saveRow, value="").fill = redFill
            ws.cell(column=19, row=saveRow, value="").fill = redFill
            ws.cell(column=2, row=saveRow, value=symbol).fill = redFill
        ws.cell(column=((len(num) - 1) * 3 + ((len(num["ROIC"]) - 1)) + 6),
                row=saveRow,value=si.get_live_price(symbol)).fill = redFill
        ws.cell(column=((len(num) - 1) * 3 + ((len(num["ROIC"]) - 1)) + 7),
                row=saveRow, value=datetime.date.today()).fill = redFill
        ws.cell(column=((len(num) - 1) * 3 + ((len(num["ROIC"]) - 1)) + 8),
                row=saveRow, value=isCash).fill = redFill
    else:
        if not isIntValueExist:
            try:
                ws.cell(column=1, row=saveRow, value=yf.Ticker(symbol).info['longName']).fill = greenFill
            except:
                pass
            ws.cell(column=2, row=saveRow, value=symbol).fill = greenFill
            ws.cell(column=18, row=saveRow, value="").fill = greenFill
            ws.cell(column=19, row=saveRow, value="").fill = greenFill
        ws.cell(column=((len(num) - 1) * 3 + ((len(num["ROIC"]) - 1)) + 6),
                row=saveRow, value=si.get_live_price(symbol)).fill = greenFill
        ws.cell(column=((len(num) - 1) * 3 + ((len(num["ROIC"]) - 1)) + 7),
                row=saveRow, value=datetime.date.today()).fill = greenFill
        ws.cell(column=((len(num) - 1) * 3 + ((len(num["ROIC"]) - 1)) + 8),
                row=saveRow, value=isCash).fill = greenFill
    for i in range(len(num)):
        for j in range(len(num["ROIC"])):
            ws.cell(column=i*3+j+3, row=saveRow, value=list(num[list(num.keys())[i]].values())[j])
    mainWB.save(filePath)

def isValidChoice(choice,ra):
    for num in range(ra):
        if choice==num+1:
            return True
    return False

def isListsValid(lt):
    for l in lt:
        if l=="Not enough Data!!":
            return False
    return True

def printUpdated():
    tk.messagebox.showinfo(title="Update Info", message="Updated was successfully done!")
    return 0

def printFailedToUpdate(symbol):
    tk.messagebox.showwarning(title="Failed to Update Stock", message=("Failed to update the Stock with symbol: " + symbol))


def updateSymbol(symbol,filePathOfExcel,isCash,isGreen):
    loc = getDataLocation()
    fileName = loc[settingBrowseNames[0]] + "\\" + symbol + " Key Ratios.xlsx"
    isDownload = os.path.isfile(fileName)==False or secToMonths(time.time()-os.stat(fileName).st_mtime)>3
    if(isDownload):
        result = downloadStockData(symbol)
        if (result == -1):
            printFailedToUpdate(symbol)
            return -1
        try:
            changePlaceForTheFile(symbol)
        except FileNotFoundError:
            messagebox.showerror("ERROR", ("Failed To Download the file Of stock with symbol: " + symbol))
            return -1
        convert_CSV_To_XLSX(symbol)
    try:
        localwb = xlrd.open_workbook(fileName)
    except FileNotFoundError:
        messagebox.showerror("ERROR","File of Symbol: " + str(symbol) + "Not Found")
        return -1
    sheetData = localwb.sheet_by_index(0)
    revenue = createList(sheetData, REVENUE)
    eps = createList(sheetData, EPS)
    equity = createList(sheetData, BOOKVALUE)
    freeCashFlow = createList(sheetData, FREECASHFLOW)
    operatingCashFlow = createList(sheetData, OPERATINGCASH)
    roic = createList(sheetData, ROIC)
    if not isListsValid([revenue, eps, equity, freeCashFlow, operatingCashFlow, roic]):
        messagebox.showerror("ERROR", ("Not enough Data To Calculate for stock with symbol: " + symbol))
        return -1
    global numbers
    numbers = {}
    numbers["ROIC"] = calNumbers(roic, "AVERAGE")
    numbers["Equity"] = calNumbers(equity, "GROWTH")
    numbers["EPS"] = calNumbers(eps, "GROWTH")
    numbers["Revenue"] = calNumbers(revenue, "GROWTH")
    if(isCash==1):
        numbers["FreeCashFlow"] = calNumbers(freeCashFlow, "GROWTH")
    else:
        numbers["OperatingCashFlow"] = calNumbers(operatingCashFlow, "GROWTH")
    if(isGreen==1):
        saveStock(filePathOfExcel, symbol, numbers, "G")
    else:
        saveStock(filePathOfExcel, symbol, numbers, "R")
    return 0

def isCashIsGreen(ICIG,sheet,numRow):
    green1 = "FF00FF00"
    green2 = "0000FF00"
    ICIG[0] = sheet.cell(numRow,22).value
    if(sheet.cell(numRow,22).fill.start_color.rgb == green1 or sheet.cell(numRow,22).fill.start_color.rgb == green2):
        ICIG[1] = 1

def updateExcel():
    global isSaveNeeded
    newWindow = tk.Toplevel(root)
    newWindow.title("Updating Excel")
    newWindow.geometry("600x100")
    my_progress = ttk.Progressbar(newWindow, orient=tk.HORIZONTAL, length=500, mode='determinate')
    my_progress.pack(pady=20)
    dataFileLoc = getDataLocation()
    excelLoc = dataFileLoc[settingBrowseNames[1]] + "\\" + "inv.xlsx"
    copyInvFile()
    isSaveNeeded=False
    wb = openpyxl.load_workbook(excelLoc)
    sheet = wb.active
    startIndex = -1
    try:
        if not sheet.cell(firstRow+1,1).value==None:
            startIndex=firstRow+1
    except:
        startIndex = -1
    if startIndex==-1:
        printUpdated()
        isSaveNeeded = True
        return
    endIndex = firstRow+1
    try:
        for i in range(3000):
            if sheet.cell(endIndex,1).value==None:
                endIndex += -1
                break
            else:
                endIndex += 1
    except:
        endIndex += -1
    ICIG = [0,0]
    retVal = 0
    jumpsVal = 100 / (endIndex - startIndex + 1)
    while startIndex<=endIndex:
        stockSymbol = sheet.cell(startIndex,2).value
        isCashIsGreen(ICIG, sheet,startIndex)
        retVal += updateSymbol(stockSymbol,excelLoc,ICIG[0],ICIG[1])
        ICIG[0] = 0
        ICIG[1] = 0
        startIndex += 1
        my_progress['value'] += jumpsVal
        my_progress.update()
    isSaveNeeded=True
    newWindow.destroy()
    if(retVal==0):
        printUpdated()
    else:
        messagebox.showinfo("Done", "Finish Updating but with some issues")

def copyInvFile():
    fileLoc = getDataLocation()
    if not os.path.isdir(os.path.join(fileLoc[settingBrowseNames[1]],"excel_copies")):
        os.mkdir(os.path.join(fileLoc[settingBrowseNames[1]],"excel_copies"))
    if isSaveNeeded and os.path.isfile(os.path.join(fileLoc[settingBrowseNames[1]],"inv.xlsx")):
        delMechanism()
        src = fileLoc[settingBrowseNames[1]] + "\\" + "inv.xlsx"
        timeStamp = datetime.datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
        dest = fileLoc[settingBrowseNames[1]] + "\\excel_copies\\" + "inv_copy_" + timeStamp +".xlsx"
        shutil.copyfile(src,dest)

def calNumbers(data,avgOrGrowth):
    dic = {}
    ROUND = 4
    if avgOrGrowth=="GROWTH":
        dic["ten"] = calGrowth(data[0], data[-2], len(data) - 2, ROUND)
        dic["five"] = calGrowth(data[-7], data[-2], 5, ROUND)
        dic["one"] = calGrowth(data[-3], data[-2], 1, ROUND)
    elif avgOrGrowth=="AVERAGE":
        dic["ten"] = calAverage(data[:-1], ROUND)
        dic["five"] = calAverage(data[-6:-1], ROUND)
        dic["one"] = calAverage(data[-2:-1], ROUND)
    return dic

def printPriceAndMos(prices):
    for i in range(6):
        if i%2==1:
            prices[i] = "MOS Price: " + str(prices[i])
        else:
            prices[i] = "Intrinsic Value: " + str(prices[i])
    rule1Header = tk.Label(calValuePage.content,text="Prices by Rule 1:")
    ben1Header = tk.Label(calValuePage.content,text="Prices by Ben Original:")
    ben2Header = tk.Label(calValuePage.content,text="Prices by Ben Update:")
    labelRule1Price = tk.Label(calValuePage.content,text=prices[0])
    labelRule1Mos = tk.Label(calValuePage.content,text=prices[1])
    labelBenOriginalPrice = tk.Label(calValuePage.content,text=prices[2])
    labelBenOriginalMos = tk.Label(calValuePage.content,text=prices[3])
    labelBenUpdatePrice = tk.Label(calValuePage.content,text=prices[4])
    labelBenUpdateMos = tk.Label(calValuePage.content,text=prices[5])
    rule1Header.grid(row="4", column="0")
    ben1Header.grid(row="6",column="0")
    ben2Header.grid(row="8",column="0")
    labelRule1Price.grid(row="5",column="0")
    labelRule1Mos.grid(row="5",column="1")
    labelBenOriginalPrice.grid(row="7",column="0")
    labelBenOriginalMos.grid(row="7",column="1")
    labelBenUpdatePrice.grid(row="9",column="0")
    labelBenUpdateMos.grid(row="9",column="1")
    globLabel.append(rule1Header)
    globLabel.append(ben1Header)
    globLabel.append(ben2Header)
    globLabel.append(labelRule1Price)
    globLabel.append(labelRule1Mos)
    globLabel.append(labelBenOriginalPrice)
    globLabel.append(labelBenOriginalMos)
    globLabel.append(labelBenUpdatePrice)
    globLabel.append(labelBenUpdateMos)


def printBig5Numbers(symbol,big5Numbers,data,pageToPrint):

    labelDataHeader = tk.Label(pageToPrint.content,text=symbol + " data:")
    label2Header = tk.Label(pageToPrint.content, text="big 5 numbers:")
    labelDataHeader.grid(row="1",column="5")
    label2Header.grid(row="10",column="5")
    globLabel.append(labelDataHeader)
    globLabel.append(label2Header)

    for j in range(1,11):
        temLabel = tk.Label(pageToPrint.content, text="YEAR_"+str(j) + "    ")
        temLabel.grid(row="2",column=j)
        globLabel.append(temLabel)

    ttmLabel = tk.Label(pageToPrint.content, text="TTM")
    ttmLabel.grid(row="2",column="11")
    globLabel.append(ttmLabel)

    i1 = 0
    for i,category in enumerate(data.keys(),3):
        tempLabel = tk.Label(pageToPrint.content, text=category+": ")
        tempLabel.grid(row=i,column="0")
        globLabel.append(tempLabel)
        for d in data[category]:
            temp2Label = tk.Label(pageToPrint.content, text=data[category][i1])
            temp2Label.grid(row=i,column=1+i1)
            globLabel.append(temp2Label)
            i1 = (i1+1)%11

    i2 = 0
    j2 = 0
    for key in numbers.keys():
        tLabel = tk.Label(pageToPrint.content, text=key + ":")
        tLabel.grid(row=12, column=i2 % 6)
        globLabel.append(tLabel)
        for key2 in numbers[key]:
            textStr=""
            if(numbers[key][key2]=="missing Data"):
                textStr="missing Data"
            else:
                textStr=key2 + ": " + str(round((numbers[key])[key2] * 100, 4))
            t2Label = tk.Label(pageToPrint.content, text=textStr)
            t2Label.grid(row=13 + j2 % 3, column=i2 % 6)
            globLabel.append(t2Label)
            j2 = j2 + 1
        i2 = i2 + 1



    # revenue = createList(sheet, REVENUE)
    # eps = createList(sheet, EPS)
    # equity = createList(sheet, BOOKVALUE)
    # freeCashFlow = createList(sheet, FREECASHFLOW)
    # operatingCashFlow = createList(sheet, OPERATINGCASH)
    # roic




# def calculateValue():
#     currentEPS = float(input("Enter current EPS: "))
#     EPSGrowth = float(input("Enter EPS growth (as percentage): "))
#     PE = float(input("Enter PE: "))
#     EPSGrowth = (EPSGrowth / 100) + 1
#     IV = PE * currentEPS * pow(EPSGrowth, 10) / 4
#     MOS = IV / 2
#     print("The Intrinsic Value is: " + str(IV))
#     print("The MOS Price is: " + str(MOS))




def getBig5Numbers(stock,pageToPrint):
    destroyTempLabel()
    switchFrames(pageToPrint,pageToPrint)
    if(not isDataLocValid()):
        messagebox.showerror("Error","Please enter valid locations inside the setting page.")
        return -1
    loc = getDataLocation()
    fileName = loc[settingBrowseNames[0]] + "\\" + stock + " Key Ratios.xlsx"
    isDownload = os.path.isfile(fileName)==False or secToMonths(time.time()-os.stat(fileName).st_mtime)>3
    if(isDownload==True):
        result = downloadStockData(stock)
        if(result==-1):
            return -1
        try:
            changePlaceForTheFile(stock)
        except FileNotFoundError:
            messagebox.showerror("ERROR", "Failed To Download the file please try again")
            return -1
        convert_CSV_To_XLSX(stock)
    try:
        wb = xlrd.open_workbook(fileName)
    except FileNotFoundError:
        messagebox.showerror("ERROR","File with the stock data not found")
        return -1
    sheet = wb.sheet_by_index(0)
    revenue = createList(sheet, REVENUE)
    eps = createList(sheet, EPS)
    equity = createList(sheet, BOOKVALUE)
    freeCashFlow = createList(sheet, FREECASHFLOW)
    operatingCashFlow = createList(sheet, OPERATINGCASH)
    roic = createList(sheet, ROIC)
    if not isListsValid([revenue, eps, equity, freeCashFlow, operatingCashFlow, roic]):
        lable = tk.Label(pageToPrint.content,text="Not Enough Data to Calculate")
        lable.grid(row=1,column=0)
        globLabel.append(lable)
        return 4
    numbers["ROIC"] = calNumbers(roic, "AVERAGE")
    numbers["Equity"] = calNumbers(equity, "GROWTH")
    numbers["EPS"] = calNumbers(eps, "GROWTH")
    numbers["Revenue"] = calNumbers(revenue, "GROWTH")
    numbers["FreeCashFlow"] = calNumbers(freeCashFlow, "GROWTH")
    numbers["OperatingCashFlow"] = calNumbers(operatingCashFlow, "GROWTH")
    printBig5Numbers(stock,numbers,{"Revenue": revenue,"EPS":eps,"Equity":equity,"FreeCash": freeCashFlow,"OperatingCash": operatingCashFlow,"ROIC":roic},pageToPrint)
    result = messagebox.askyesno("SaveData","Do you want to save the stock?")
    if(result==True):
        fileLoc = getDataLocation()
        saveTo = fileLoc[settingBrowseNames[1]]+"\\" + "inv.xlsx"
        res2 = messagebox.askyesno("SaveData","Do you want to save as green?",)
        res3 = messagebox.askyesno("SaveData","Do you want to save the cash?")
        if(res3==True):
            del numbers["OperatingCashFlow"]
        else:
            del numbers["FreeCashFlow"]
        if (res2 == True):
            saveStock(saveTo, stock, numbers, "G")
        else:
            saveStock(saveTo, stock, numbers, "R")
    return 1



def convert_CSV_To_XLSX(stock):
    loc = getDataLocation()
    oldname = stock + " Key Ratios.csv"
    oldPath = loc[settingBrowseNames[0]] + "\\" + oldname
    newname = stock + " Key Ratios.xlsx"
    newPath = loc[settingBrowseNames[0]] + "\\" + newname
    if (os.path.isfile(newPath)==True):
        os.remove(newPath)
    wbook = Workbook(newPath)
    sheet1 = wbook.add_worksheet()
    count = 0
    with open(oldPath, 'rt') as f:
        data = csv.reader(f)
        for row in data:
            if (len(row) <= 0):
                count = count + 1
                continue
            if row[-1] != "TTM" and not row[-1].startswith("Latest"):
                for i in range(len(row)):
                    if i > 0 and row[i] != "":
                        row[i] = row[i].replace(",", "")
                        row[i] = float(row[i])
            sheet1.write_row(count, 0, row)
            count = count + 1
    wbook.close()


def downloadStockData(stock):
    drive = wb.Chrome()
    stock = stock.upper()
    drive.implicitly_wait(4)
    drive.get("https://financials.morningstar.com/ratios/r.html?t=" + stock)
    try:
        drive.find_element_by_css_selector('.large_button').click()
    except exceptions.NoSuchElementException:
        messagebox.showerror("ERROR!","No Such Stock Exists")
        return -1
    filename = stock + " Key Ratios.csv"
    userName = getpass.getuser()
    src = "C:\\Users\\" + userName + "\\Downloads\\" + filename
    while not os.path.isfile(src):
        pass
    drive.quit()
    return 1

def changePlaceForTheFile(stock):
    loc=getDataLocation()
    filename = stock + " Key Ratios.csv"
    userName = getpass.getuser()
    src = "C:\\Users\\" + userName + "\\Downloads\\" + filename
    dest = loc[settingBrowseNames[0]] + "\\" +  filename
    shutil.move(src,dest)


def calVal(epsGrowth,PE,currentEPS):
    destroyTempLabel()
    isGood = checkCalculatorInput({epsGrowth,PE,currentEPS})
    if not isGood:
        messagebox.showerror("Erron","Invalid entries, Please insert valid numbers")
        return
    epsGrowth = float(epsGrowth)
    PE = float(PE)
    currentEPS = float(currentEPS)
    rule1Price = rule1Calculator((epsGrowth/100)+1,PE,currentEPS)
    rule1Mos = rule1Price/2
    ben1Price = benGrahamOriginal(epsGrowth,currentEPS)
    ben1Mos = ben1Price/2
    ben2Price = benGrahamUpdate(epsGrowth,currentEPS)
    ben2Mos = ben2Price/2
    prices= [rule1Price,rule1Mos,ben1Price,ben1Mos,ben2Price,ben2Mos]
    printPriceAndMos(prices)


    # currentEPS = float(input("Enter current EPS: "))
    # EPSGrowth = float(input("Enter EPS growth (as percentage): "))
    # PE = float(input("Enter PE: "))
    # EPSGrowth = (EPSGrowth / 100) + 1
    # IV = PE * currentEPS * pow(EPSGrowth, 10) / 4
    # MOS = IV / 2
    # print("The Intrinsic Value is: " + str(IV))
    # print("The MOS Price is: " + str(MOS))

def autoStocksFromList():
    stockListFile = open("stockList.txt","r")
    listOfStock = []
    for line in stockListFile:
        if not line.startswith("#"):
            listOfStock.append(line.rstrip())
    stockListFile.close()
    loc = getDataLocation()
    for stock in listOfStock:
        fileName = loc[settingBrowseNames[0]] + "\\" + stock + " Key Ratios.xlsx"
        isDownload = os.path.isfile(fileName) == False or secToMonths(time.time() - os.stat(fileName).st_mtime) > 3
        if(isDownload):
            result = downloadStockData(stock)
            if (result == -1):
                messagebox.showerror("ERROR", "Failed To Download the file please try again")
                return None
            try:
                changePlaceForTheFile(stock)
            except FileNotFoundError:
                messagebox.showerror("ERROR", "Failed To Download the file please try again")
                return None
            convert_CSV_To_XLSX(stock)
    toContinue = messagebox.askyesno("To Show Data","Finished Downloading, Do you want continue to read the Data?")
    notEnoughtDataList = []
    if(toContinue):
        for stock in listOfStock:
            ans = getBig5Numbers(stock, autoPage)
            if(ans==-1):
                return None
            elif(ans==4):
                messagebox.showinfo("Update","Not enough Data To Calculate for stock with symbol: "+ stock)
                notEnoughtDataList.append(stock)
    return notEnoughtDataList

def cleanUpAutoPage():
    switchFrames(autoPage,mainPage)

###########################################################
###############        mainPage           #################
###########################################################

option1 = tk.Button(mainPage.content, text="Calculate Stock Data", width=40, command=lambda:switchFrames(mainPage,stockDataPage))
option2 = tk.Button(mainPage.content, text="Calculate intrinsic value", width=40, command=lambda:switchFrames(mainPage,calValuePage))
option3 = tk.Button(mainPage.content,text="Settings", width=40, command=lambda:switchFrames(mainPage,settingsPage))
option4 = tk.Button(mainPage.content, text="About", width=40, command=lambda:switchFrames(mainPage,aboutPage))
option5 = tk.Button(mainPage.content, text="Automate", width=40, command=lambda:switchFrames(mainPage,autoPage))
option6 = tk.Button(mainPage.content, text="Update Excel", width=40, command=updateExcel)
option7 = tk.Button(mainPage.content, text="Exit", width=40, command=root.quit)
option1.grid(row=0,   column=0, pady=7)
option2.grid(row=1,   column=0, pady=7)
option3.grid(row=2,   column=0, pady=7)
option4.grid(row=3,   column=0, pady=7)
option5.grid(row=4,   column=0, pady=7)
option6.grid(row=5,   column=0, pady=7)
option7.grid(row=100, column=0, pady=7)
mainPage.show()



###########################################################
###############        StockDataPage      #################
###########################################################

dataButton = tk.Button(stockDataPage.content, text="Get Stock Data", command=lambda: getBig5Numbers(str(symbol.get()).upper(),stockDataPage))
back = tk.Button(stockDataPage.content, text="Back",command=lambda:switchFrames(stockDataPage,mainPage))
symbol = tk.Entry(stockDataPage.content, width=20, borderwidth=5)
symbol.bind('<FocusIn>', onFocusEntry)
symbol.bind('<FocusOut>', lambda event, param="Enter Symbol": onFocusOut(event,param))
symbol.grid(row=0,column=0)
dataButton.grid(row=0,column=1)
back.grid(row=100,column=0,pady=10)

###########################################################
###############        autoPage           #################
###########################################################

startAutoMate = tk.Button(autoPage.content, text="start", command=autoStocksFromList)
startAutoMate.grid(row=0,column=6, pady=7)
backButton = tk.Button(autoPage.content, text="back", command=cleanUpAutoPage)
backButton.grid(row=100, column=6, pady=7)

###########################################################
###############     calValuePage          #################
###########################################################

back = tk.Button(calValuePage.content, text="Back", command=lambda: switchFrames(calValuePage,mainPage))
entryValEps = tk.Entry(calValuePage.content, width=30, borderwidth=5)
entryValEps.bind('<FocusIn>',onFocusEntry)
entryValEps.bind('<FocusOut>',lambda event, param="EPS...": onFocusOut(event,param))
entryValEps.grid(row=0, column=1)
labelValEps = tk.Label(calValuePage.content, text="Current EPS: ")
labelValEps.grid(row=0, column=0)
entryGrowthEps = tk.Entry(calValuePage.content, width=30, borderwidth=5)
entryGrowthEps.bind('<FocusIn>',onFocusEntry)
entryGrowthEps.bind('<FocusOut>',lambda event, param="EPS Growth...": onFocusOut(event,param))
entryGrowthEps.grid(row=1, column=1)
labelGrowthEps = tk.Label(calValuePage.content, text="EPS Growth: ")
labelGrowthEps.grid(row=1, column=0)
entryPe = tk.Entry(calValuePage.content, width=30, borderwidth=5)
entryPe.bind('<FocusIn>',onFocusEntry)
entryPe.bind('<FocusOut>',lambda event, param="PE...": onFocusOut(event,param))
entryPe.grid(row=2, column=1)
labelPe = tk.Label(calValuePage.content, text="Forword PE: ")
labelPe.grid(row=2, column=0)
bCal = tk.Button(calValuePage.content, text="Calculate Value",command=lambda: calVal(entryGrowthEps.get(),entryPe.get(),entryValEps.get()))
bCal.grid(row=3, column=0, pady=5)
back.grid(row=100, column=0)


###########################################################
###############     SettingPage           #################
###########################################################

backSetting = tk.Button(settingsPage.content, text="Back", command=lambda: switchFrames(settingsPage,mainPage))
stockFilesEntry = tk.Entry(settingsPage.content, width="60", borderwidth="5")
stockFileLabel = tk.Label(settingsPage.content, text="stock data files location: ")
stockFilesButton = tk.Button(settingsPage.content, text="browse", command=lambda: browseLocation(settingBrowseNames[0]))
favStocksEntry = tk.Entry(settingsPage.content, width="60", borderwidth="5")
favStocksLabel = tk.Label(settingsPage.content, text="favorite stocks location: ")
favStocksButton = tk.Button(settingsPage.content, text="browse", command=lambda: browseLocation(settingBrowseNames[1]))
saveButton = tk.Button(settingsPage.content, text="save",command=lambda: saveLocToDataFileWrapper([stockFilesEntry.get(),favStocksEntry.get()]),state=tk.DISABLED)
writeFromData()
stockFilesEntry.bind('<KeyPress>',onChangeLoc)
favStocksEntry.bind('<KeyPress>',onChangeLoc)
stockFileLabel.grid(row=0,column=0)
stockFilesEntry.grid(row=0,column=1)
stockFilesButton.grid(row=0,column=2)
favStocksLabel.grid(row=1,column=0)
favStocksEntry.grid(row=1,column=1)
favStocksButton.grid(row=1,column=2)
backSetting.grid(row=100, column=0)
saveButton.grid(row=100, column=1)
# TODO: Build The setting page
# TODO: If he change location then I should move the files to the new location.

###########################################################
###############     AboutPage             #################
###########################################################

# TODO: Build the AboutPage


###########################################################
###############     AutoMate              #################
###########################################################



##########################################################
#####      main program:                 #################
##########################################################


root.mainloop()